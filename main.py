import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart,Reference
from openpyxl.utils import get_column_letter

df = pd.read_excel('supermarket_sales.xlsx')
df = df[['Gender','Product line','Total']]

pivot_table = df.pivot_table(index='Gender',columns='Product line',values='Total',aggfunc='sum').round(0)
pivot_table.to_excel('pivot_table.xlsx',sheet_name='Report',startrow=4)

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

#sum of totals in the columns in our pivot table
for i in range(min_column+1,max_column+1):
    letter = get_column_letter(i) #gets the letter based on the number, eg 1 is col A
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = "Currency"
wb.save('report.xlsx')

barchart = BarChart()
data = Reference(sheet,
          min_col = min_column+1, #data starts in col B, not A so we add 1
          max_col = max_column,
          min_row = min_row,
          max_row = max_row)

categories = Reference(sheet,
          min_col = min_column, #the 2 genders are in the same column A (male and female)
          max_col = min_column,
          min_row = min_row+1, #move from header to data row by adding 1
          max_row = max_row)

#add data and categories
barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categories)

#indicate where we want the barchart placed
sheet.add_chart(barchart,"B12")

barchart.title = "Sales by Product Line"
barchart.style = 3
wb.save('barchart.xlsx')